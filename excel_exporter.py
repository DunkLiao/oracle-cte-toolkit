from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


INVALID_SHEET_CHARS = (":", "\\", "/", "?", "*", "[", "]")


def sanitize_sheet_name(name):
    cleaned = str(name)
    for char in INVALID_SHEET_CHARS:
        cleaned = cleaned.replace(char, "")
    cleaned = cleaned.strip()
    return (cleaned or "Sheet")[:31]


def add_dataframe_sheet(workbook, sheet_name, dataframe):
    worksheet = workbook.create_sheet(title=sanitize_sheet_name(sheet_name))
    for row in dataframe_to_rows(dataframe, index=False, header=True):
        worksheet.append(row)
    return worksheet


def export_query_dataframes(dataframes, output_path):
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, dataframe in dataframes.items():
        add_dataframe_sheet(workbook, sheet_name, dataframe)

    if not workbook.sheetnames:
        workbook.create_sheet(title="EMPTY")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def export_metadata_rows(rows, output_path):
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    dataframe = pd.DataFrame(rows)
    dataframe.to_excel(output_path, index=False, engine="openpyxl")
