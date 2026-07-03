import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from excel_exporter import (
    export_metadata_rows,
    export_query_dataframes,
    sanitize_sheet_name,
)


class ExcelExporterTests(unittest.TestCase):
    def test_sanitize_sheet_name_removes_invalid_characters_and_truncates(self):
        name = sanitize_sheet_name("abc:def/ghi*jkl[mno]01234567890123456789")

        self.assertNotIn(":", name)
        self.assertNotIn("/", name)
        self.assertLessEqual(len(name), 31)

    def test_export_query_dataframes_creates_empty_sheet_when_no_results(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "out.xlsx"

            export_query_dataframes({}, path)
            wb = load_workbook(path)

        self.assertEqual(wb.sheetnames, ["EMPTY"])

    def test_export_query_dataframes_writes_dataframe_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "out.xlsx"
            frames = {"sample": pd.DataFrame([{"A": 1, "B": "x"}])}

            export_query_dataframes(frames, path)
            ws = load_workbook(path)["sample"]

        self.assertEqual(ws["A1"].value, "A")
        self.assertEqual(ws["B2"].value, "x")

    def test_export_metadata_rows_writes_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "metadata.xlsx"

            export_metadata_rows(
                [{"SQL_FILE": "a.sql", "COLUMN_NAME": "COL1"}],
                path,
            )
            ws = load_workbook(path).active

        self.assertEqual(ws["A1"].value, "SQL_FILE")
        self.assertEqual(ws["B2"].value, "COL1")


if __name__ == "__main__":
    unittest.main()
